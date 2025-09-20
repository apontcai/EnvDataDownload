import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
import os
from pathlib import Path
import threading


class ExcelProcessor:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Data Processor")
        self.root.geometry("600x400")

        # Variables
        self.folder_path = tk.StringVar()

        self.create_widgets()

    def create_widgets(self):
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Folder selection
        ttk.Label(main_frame, text="Select Folder:").grid(row=0, column=0, sticky=tk.W, pady=5)

        folder_frame = ttk.Frame(main_frame)
        folder_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)

        ttk.Entry(folder_frame, textvariable=self.folder_path, width=50).grid(row=0, column=0, sticky=(tk.W, tk.E))
        ttk.Button(folder_frame, text="Browse", command=self.browse_folder).grid(row=0, column=1, padx=(5, 0))

        folder_frame.columnconfigure(0, weight=1)

        # Process button
        ttk.Button(main_frame, text="Process Files", command=self.start_processing).grid(row=2, column=0, pady=20)

        # Progress bar
        self.progress = ttk.Progressbar(main_frame, length=400, mode='determinate')
        self.progress.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)

        # Status label
        self.status_label = ttk.Label(main_frame, text="Ready to process files")
        self.status_label.grid(row=4, column=0, columnspan=2, pady=5)

        # Log text area
        ttk.Label(main_frame, text="Processing Log:").grid(row=5, column=0, sticky=tk.W, pady=(10, 0))

        log_frame = ttk.Frame(main_frame)
        log_frame.grid(row=6, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)

        self.log_text = tk.Text(log_frame, height=15, width=70)
        scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)

        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))

        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)

        # Configure grid weights
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(6, weight=1)
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

    def browse_folder(self):
        folder_selected = filedialog.askdirectory()
        if folder_selected:
            self.folder_path.set(folder_selected)

    def log_message(self, message):
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def start_processing(self):
        if not self.folder_path.get():
            messagebox.showerror("Error", "Please select a folder first!")
            return

        # Run processing in a separate thread to prevent UI freezing
        thread = threading.Thread(target=self.process_files)
        thread.daemon = True
        thread.start()

    def process_files(self):
        try:
            folder = Path(self.folder_path.get())

            # Clear log
            self.log_text.delete(1.0, tk.END)

            # Find template file
            template_file = self.find_template_file(folder)
            if not template_file:
                self.log_message("ERROR: No template file found (should start with 'template')")
                return

            self.log_message(f"Found template file: {template_file.name}")

            # Check template file structure
            if not self.validate_template_file(template_file):
                return

            # Find raw data files (Excel files that don't start with 'template')
            raw_files = self.find_raw_files(folder)
            if not raw_files:
                self.log_message("ERROR: No raw data files found")
                return

            self.log_message(f"Found {len(raw_files)} raw data files to process")

            # Setup progress bar
            self.progress['maximum'] = len(raw_files)
            self.progress['value'] = 0

            # Process each raw file
            for i, raw_file in enumerate(raw_files):
                self.status_label.config(text=f"Processing {raw_file.name}...")
                self.log_message(f"\nProcessing: {raw_file.name}")

                try:
                    self.process_single_file(template_file, raw_file)
                    self.log_message(f"✓ Successfully processed: {raw_file.name}")
                except Exception as e:
                    self.log_message(f"✗ Error processing {raw_file.name}: {str(e)}")

                self.progress['value'] = i + 1
                self.root.update_idletasks()

            self.status_label.config(text="Processing completed!")
            self.log_message(f"\n=== Processing completed! ===")
            messagebox.showinfo("Success", "All files have been processed!")

        except Exception as e:
            self.log_message(f"ERROR: {str(e)}")
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    def find_template_file(self, folder):
        """Find the first Excel file that starts with 'template'"""
        for file in folder.glob("template*.xlsx"):
            return file
        # Also try .xls files
        for file in folder.glob("template*.xls"):
            return file
        return None

    def find_raw_files(self, folder):
        """Find all Excel files that don't start with 'template'"""
        raw_files = []
        # Check both .xlsx and .xls files
        for pattern in ["*.xlsx", "*.xls"]:
            for file in folder.glob(pattern):
                if not file.name.lower().startswith('template') and not file.name.endswith(
                        '_processed.xlsx') and not file.name.endswith('_processed.xls'):
                    raw_files.append(file)
        return raw_files

    def validate_template_file(self, template_file):
        """Validate template file and show available sheets"""
        try:
            wb = openpyxl.load_workbook(template_file)
            sheet_names = wb.sheetnames

            self.log_message(f"Template worksheets: {sheet_names}")

            # Try to find the data sheet
            data_sheet = self.find_data_sheet(wb)
            if data_sheet:
                self.log_message(f"Using template sheet: '{data_sheet.title}' for data")
                wb.close()
                return True
            else:
                self.log_message("ERROR: Could not find a suitable data sheet in template")
                wb.close()
                return False

        except Exception as e:
            self.log_message(f"ERROR: Could not validate template file: {str(e)}")
            return False

    def find_data_sheet(self, workbook):
        """Find the data sheet in the workbook"""
        sheet_names = workbook.sheetnames

        # First, try to find a sheet named "data" (case-insensitive)
        for name in sheet_names:
            if name.lower() == 'data':
                return workbook[name]

        # If no "data" sheet found, try other common names
        data_keywords = ['data', 'monitoring', 'readings', 'measurements', 'values']
        for name in sheet_names:
            for keyword in data_keywords:
                if keyword in name.lower():
                    return workbook[name]

        # If still not found, use the first sheet
        if sheet_names:
            return workbook[sheet_names[0]]

        return None

    def process_single_file(self, template_file, raw_file):
        """Process a single raw data file using the template"""

        # Load the template workbook
        template_wb = openpyxl.load_workbook(template_file)

        # Find the data sheet in template
        data_sheet = self.find_data_sheet(template_wb)
        if not data_sheet:
            raise Exception("Could not find data sheet in template")

        # Load the raw data workbook - just get the first/only sheet
        raw_wb = openpyxl.load_workbook(raw_file)
        raw_sheet = raw_wb.worksheets[0]  # Get the first (and only) sheet

        self.log_message(f"  → Using template sheet: '{data_sheet.title}'")
        self.log_message(f"  → Using raw data sheet: '{raw_sheet.title}'")

        # Clear existing data in template (columns A and B from row 2 onwards)
        self.clear_columns(data_sheet, ['A', 'B'], start_row=2)

        # Copy data from raw file to template
        rows_copied = self.copy_data(raw_sheet, data_sheet)
        self.log_message(f"  → Copied {rows_copied} rows of data")

        # Update chart title if chart exists
        self.update_chart_title(template_wb, raw_file.stem)

        # Save the processed file
        output_filename = raw_file.stem + "_processed.xlsx"
        output_path = raw_file.parent / output_filename
        template_wb.save(output_path)

        # Close workbooks
        template_wb.close()
        raw_wb.close()

        self.log_message(f"  → Saved as: {output_filename}")

    def clear_columns(self, sheet, columns, start_row=2):
        """Clear specified columns from start_row to the end"""
        max_row = sheet.max_row

        for col in columns:
            for row in range(start_row, max_row + 100):  # Clear extra rows to be safe
                sheet[f"{col}{row}"] = None

    def copy_data(self, source_sheet, target_sheet):
        """Copy data from source sheet columns A and B to target sheet"""

        # Find the last row with data in source sheet
        max_row = source_sheet.max_row
        rows_copied = 0

        # Copy column A (time) and B (reading) from row 2 onwards
        for row in range(2, max_row + 1):
            time_value = source_sheet[f"A{row}"].value
            reading_value = source_sheet[f"B{row}"].value

            # Only copy if there's actual data (at least one value is not None)
            if time_value is not None or reading_value is not None:
                target_sheet[f"A{row}"] = time_value
                target_sheet[f"B{row}"] = reading_value
                rows_copied += 1
            elif rows_copied > 0:
                # If we've already copied some data and hit empty rows, stop
                break

        return rows_copied

    def update_chart_title(self, workbook, new_title):
        """Update chart title in the workbook - simplified approach"""
        try:
            charts_updated = 0

            # Check all worksheets for embedded charts
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                if hasattr(sheet, '_charts') and sheet._charts:
                    self.log_message(f"  → Found {len(sheet._charts)} chart(s) in sheet '{sheet_name}'")

                    for chart_idx, chart in enumerate(sheet._charts):
                        try:
                            self.log_message(f"  → Processing chart {chart_idx + 1} in sheet '{sheet_name}'...")

                            # Get current title for debugging
                            current_title = self.get_current_chart_title(chart)
                            self.log_message(f"  → Current title: '{current_title}'")

                            # Try the simplest approach first - direct string assignment
                            success = False

                            # Method 1: Direct string assignment (simplest)
                            try:
                                chart.title = new_title
                                success = True
                                self.log_message(f"  → Method 1 (direct assignment) successful")
                            except Exception as e1:
                                self.log_message(f"  → Method 1 failed: {str(e1)}")

                            # Method 2: Create proper Title object
                            if not success:
                                try:
                                    from openpyxl.chart.title import Title
                                    title_obj = Title()
                                    title_obj.text = new_title
                                    chart.title = title_obj
                                    success = True
                                    self.log_message(f"  → Method 2 (Title object) successful")
                                except Exception as e2:
                                    self.log_message(f"  → Method 2 failed: {str(e2)}")

                            # Method 3: Create Title with Rich Text
                            if not success:
                                try:
                                    from openpyxl.chart.title import Title
                                    from openpyxl.chart.text import RichText
                                    from openpyxl.drawing.text import Paragraph, ParagraphProperties, \
                                        CharacterProperties, Run

                                    # Create rich text structure
                                    title_obj = Title()
                                    rich_text = RichText()
                                    paragraph = Paragraph()
                                    run = Run()
                                    run.t = new_title
                                    paragraph.r.append(run)
                                    rich_text.p.append(paragraph)
                                    title_obj.tx = rich_text

                                    chart.title = title_obj
                                    success = True
                                    self.log_message(f"  → Method 3 (Rich Text) successful")
                                except Exception as e3:
                                    self.log_message(f"  → Method 3 failed: {str(e3)}")

                            if success:
                                charts_updated += 1
                                self.log_message(
                                    f"  → Successfully updated chart {chart_idx + 1} title to: '{new_title}'")

                                # Verify the title was set
                                new_current_title = self.get_current_chart_title(chart)
                                self.log_message(f"  → Verified new title: '{new_current_title}'")
                            else:
                                self.log_message(f"  → All methods failed for chart {chart_idx + 1}")

                        except Exception as chart_error:
                            self.log_message(f"  → Error processing chart {chart_idx + 1}: {str(chart_error)}")

            if charts_updated == 0:
                self.log_message("  → No charts were successfully updated")
            else:
                self.log_message(f"  → Successfully updated {charts_updated} chart(s)")

        except Exception as e:
            self.log_message(f"  → Error in chart title update process: {str(e)}")

    def get_current_chart_title(self, chart):
        """Get the current chart title for debugging"""
        try:
            if hasattr(chart, 'title') and chart.title is not None:
                if hasattr(chart.title, 'text') and chart.title.text is not None:
                    return chart.title.text
                elif hasattr(chart.title, 'tx') and chart.title.tx is not None:
                    if hasattr(chart.title.tx, 'text'):
                        return chart.title.tx.text
                    elif hasattr(chart.title.tx, 'rich') and chart.title.tx.rich is not None:
                        if hasattr(chart.title.tx.rich, 'p') and chart.title.tx.rich.p:
                            for p in chart.title.tx.rich.p:
                                if hasattr(p, 'r') and p.r:
                                    for r in p.r:
                                        if hasattr(r, 't'):
                                            return r.t
                return "Title object exists but no text found"
            return "No title object"
        except Exception as e:
            return f"Error reading title: {str(e)}"


def main():
    root = tk.Tk()
    app = ExcelProcessor(root)
    root.mainloop()


if __name__ == "__main__":
    main()